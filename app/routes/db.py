"""Define rutas y utilidades para empleados, asistencias, prenómina y relojes biométricos."""

from fastapi import APIRouter, HTTPException, Query
from datetime import datetime, time, timedelta
from pydantic import BaseModel
from typing import Optional, Literal
from fastapi.responses import StreamingResponse
from io import BytesIO
import json
import socket

from app.services.db_service import DBService
from app.services.zk_service import ZKService
from app.exceptions import DeviceClockDriftError
from app.utils.response import success
from app.services.excel_service import build_attendance_excel

try:
    from app.services.excel_service import build_payroll_excel
except ImportError:
    build_payroll_excel = None


router = APIRouter(
    prefix="/db",
    tags=["Base de Datos"]
)

DATE_FORMATS = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y")
EmpresaDevice = Literal["FISMAN", "SELEFF"]
WEEKDAY_LABELS = [
    "LUNES",
    "MARTES",
    "MIERCOLES",
    "JUEVES",
    "VIERNES",
    "SABADO",
    "DOMINGO",
]


def parse_date(value: str):
    """Convierte una cadena de texto en una fecha válida usando los formatos admitidos.

    :param value: Cadena que contiene la fecha que se desea interpretar.
    :type value: str
    :return: Fecha obtenida a partir de la cadena proporcionada.
    :rtype: datetime.date
    :raises HTTPException: Si la cadena no coincide con ninguno de los formatos admitidos.
    """
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            continue

    raise HTTPException(
        status_code=400,
        detail="Formato de fecha inválido. Usa YYYY-MM-DD, DD-MM-YYYY o DD/MM/YYYY"
    )


def parse_hour(value: str):
    """Convierte una cadena de texto en una hora con formato de horas y minutos.

    :param value: Cadena que contiene la hora que se desea interpretar.
    :type value: str
    :return: Hora obtenida a partir de la cadena proporcionada.
    :rtype: datetime.time
    :raises HTTPException: Si la cadena no utiliza el formato HH:MM.
    """
    try:
        return datetime.strptime(value.strip(), "%H:%M").time()
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Formato de hora inválido. Usa HH:MM, ejemplo 15:00"
        )


def parse_user_ids(value: Optional[str]):
    """Separa una cadena de identificadores de empleados delimitados por comas.

    :param value: Cadena de identificadores o valor nulo.
    :type value: Optional[str]
    :return: Lista de identificadores sin espacios vacíos.
    :rtype: list[str]
    """
    if not value:
        return []

    return [
        item.strip()
        for item in value.split(",")
        if item.strip()
    ]


def make_assignment_key(device_id, user_id) -> str:
    """Crea una clave única a partir del reloj y el código de empleado.

    :param device_id: Identificador del reloj asociado.
    :type device_id: int | None
    :param user_id: Código del empleado asociado.
    :type user_id: str | int | None
    :return: Clave compuesta con el formato ``device_id:user_id``.
    :rtype: str
    """
    clean_device_id = int(device_id) if device_id is not None else 0
    return f"{clean_device_id}:{str(user_id or '').strip()}"


def parse_assignment_key(value: str):
    """Interpreta una clave compuesta de reloj y empleado.

    :param value: Clave que se desea interpretar.
    :type value: str
    :return: Tupla con el identificador del reloj y el código del empleado, o ``None`` si la clave no es válida.
    :rtype: tuple[int, str] | None
    """
    raw = str(value or "").strip()
    if ":" not in raw:
        return None

    device_raw, user_id = raw.split(":", 1)
    try:
        device_id = int(device_raw)
    except (TypeError, ValueError):
        return None

    user_id = user_id.strip()
    if not user_id:
        return None

    return device_id, user_id


def matches_assignment(entity, selected_values: set[str]) -> bool:
    """Comprueba si una entidad coincide con alguna asignación seleccionada.

    :param entity: Entidad que contiene datos de reloj, usuario o UID.
    :type entity: object
    :param selected_values: Claves compuestas o identificadores simples seleccionados.
    :type selected_values: set[str]
    :return: ``True`` si la entidad coincide o si no existen filtros; en caso contrario, ``False``.
    :rtype: bool
    """
    if not selected_values:
        return True

    entity_device_id = getattr(entity, "device_id", None)
    entity_user_id = str(getattr(entity, "user_id", "") or "").strip()
    entity_uid = str(getattr(entity, "uid", "") or "").strip()

    for selected in selected_values:
        parsed = parse_assignment_key(selected)
        if parsed:
            device_id, user_id = parsed
            if (
                entity_device_id == device_id
                and user_id in {entity_user_id, entity_uid}
            ):
                return True
        elif selected in {entity_user_id, entity_uid}:
            return True

    return False


def get_date_range_days(start_date: str, end_date: str):
    """Obtiene todas las fechas comprendidas dentro de un intervalo inclusivo.

    :param start_date: Fecha inicial en uno de los formatos admitidos.
    :type start_date: str
    :param end_date: Fecha final en uno de los formatos admitidos.
    :type end_date: str
    :return: Lista ordenada de fechas desde el inicio hasta el final.
    :rtype: list[datetime.date]
    :raises HTTPException: Si el formato es inválido o la fecha inicial es posterior a la fecha final.
    """
    start = parse_date(start_date)
    end = parse_date(end_date)

    if start > end:
        raise HTTPException(
            status_code=400,
            detail="La fecha inicial no puede ser mayor que la fecha final"
        )

    total_days = (end - start).days

    return [start + timedelta(days=index) for index in range(total_days + 1)]


def format_day_label(day):
    """Genera la etiqueta legible de una fecha con el nombre del día de la semana.

    :param day: Fecha que se desea representar.
    :type day: datetime.date
    :return: Etiqueta con el día de la semana y la fecha.
    :rtype: str
    """
    weekday = WEEKDAY_LABELS[day.weekday()]
    return f"{weekday} {day.strftime('%d/%m/%Y')}"


def incident_to_dict(incident):
    """Convierte una incidencia de prenómina en un diccionario serializable.

    :param incident: Incidencia que se desea convertir.
    :type incident: object
    :return: Diccionario con los datos disponibles de la incidencia.
    :rtype: dict
    """
    if hasattr(incident, "to_dict"):
        return incident.to_dict()

    return {
        "id": incident.id,
        "uid": getattr(incident, "uid", None),
        "device_id": getattr(incident, "device_id", None),
        "assignment_key": make_assignment_key(
            getattr(incident, "device_id", None),
            incident.user_id,
        ),
        "user_id": incident.user_id,
        "fecha": incident.fecha.isoformat() if incident.fecha else None,
        # Compatibilidad con el frontend: ya no existe una columna `hora`
        # ni una columna `descripcion` en payroll_incidents.
        "hora": None,
        "incidencia": incident.incidencia,
        "descripcion": None,
        "color": getattr(incident, "color", None) or "#BAE6FD",
        "source_fecha": incident.source_fecha.isoformat() if getattr(incident, "source_fecha", None) else None,
        "source_hora": incident.source_hora.strftime("%H:%M") if getattr(incident, "source_hora", None) else None,
        "moved_attendance": getattr(incident, "moved_attendance", None),
        "created_at": incident.created_at.isoformat() if getattr(incident, "created_at", None) else None,
        "updated_at": incident.updated_at.isoformat() if getattr(incident, "updated_at", None) else None,
    }


def build_payroll_data(
    start_date: str,
    end_date: str,
    branch_id: Optional[int] = None,
    user_ids: Optional[list[str]] = None,
):
    """Genera la información de prenómina separando cada asignación por reloj y empleado.

    :param start_date: Fecha inicial del periodo.
    :type start_date: str
    :param end_date: Fecha final del periodo.
    :type end_date: str
    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :param user_ids: Claves o códigos opcionales de empleados seleccionados.
    :type user_ids: Optional[list[str]]
    :return: Estructura con días, horas y filas de asistencias e incidencias.
    :rtype: dict
    :raises HTTPException: Si el intervalo, la sucursal o los servicios requeridos no son válidos.
    """
    selected_assignments = set(user_ids or [])

    days = get_date_range_days(start_date, end_date)
    start_datetime = datetime.combine(days[0], time.min)
    end_datetime = datetime.combine(days[-1], time.max)

    if branch_id is not None:
        get_branch_or_404(branch_id)
        users = DBService.get_users_by_branch(branch_id)
    else:
        users = DBService.get_all_users_from_db()

    users = [
        user
        for user in users
        if str(getattr(user, "status", "Activo")).lower() == "activo"
    ]

    if selected_assignments:
        users = [
            user
            for user in users
            if matches_assignment(user, selected_assignments)
        ]

    records = get_attendance_records(
        start_datetime=start_datetime,
        end_datetime=end_datetime,
        branch_id=branch_id,
    )

    if selected_assignments:
        records = [
            record
            for record in records
            if matches_assignment(record, selected_assignments)
        ]

    try:
        incidents = DBService.get_payroll_incidents_by_range(
            start_date=days[0],
            end_date=days[-1],
            branch_id=branch_id,
        )
    except AttributeError:
        raise HTTPException(
            status_code=500,
            detail="Falta DBService.get_payroll_incidents_by_range en el backend",
        )

    if selected_assignments:
        incidents = [
            incident
            for incident in incidents
            if matches_assignment(incident, selected_assignments)
        ]

    user_assignments_by_code = {}
    for user in users:
        code = str(
            getattr(user, "user_id", None)
            or getattr(user, "uid", None)
            or ""
        ).strip()
        if not code:
            continue
        user_assignments_by_code.setdefault(code, []).append(user)

    attendance_by_assignment_date_hour = {}

    for record in records:
        if not record.timestamp:
            continue

        record_user_id = str(
            getattr(record, "user_id", None)
            or getattr(record, "uid", None)
            or ""
        ).strip()
        if not record_user_id:
            continue

        assignment_key = make_assignment_key(
            getattr(record, "device_id", None),
            record_user_id,
        )
        hour_key = f"{record.timestamp.hour:02d}:00"
        date_key = record.timestamp.date().isoformat()
        map_key = (assignment_key, date_key, hour_key)
        value = record.timestamp.strftime("%d/%m/%Y %H:%M")

        attendance_by_assignment_date_hour.setdefault(map_key, []).append(value)

    incidents_by_assignment_date_hour = {}

    for incident in incidents:
        incident_user_id = str(
            getattr(incident, "user_id", "") or ""
        ).strip()
        incident_device_id = getattr(incident, "device_id", None)

        # Compatibilidad con incidencias antiguas sin device_id: sólo se
        # recuperan automáticamente cuando el código pertenece a un único reloj.
        if incident_device_id is None:
            candidates = user_assignments_by_code.get(incident_user_id, [])
            if len(candidates) == 1:
                incident_device_id = getattr(candidates[0], "device_id", None)

        assignment_key = make_assignment_key(
            incident_device_id,
            incident_user_id,
        )
        # La incidencia ya no tiene una hora propia. Para conservar la
        # estructura horaria de la prenómina se usa source_hora cuando existe;
        # de lo contrario se ancla a la primera fila (06:00).
        incident_hour = getattr(incident, "source_hora", None) or time(6, 0)
        hour_key = f"{incident_hour.hour:02d}:00"
        date_key = incident.fecha.isoformat()
        map_key = (assignment_key, date_key, hour_key)
        incidents_by_assignment_date_hour.setdefault(map_key, []).append(incident)

    hour_labels = [f"{hour:02d}:00" for hour in range(6, 19)]
    day_payload = [
        {
            "date": day.isoformat(),
            "label": format_day_label(day),
        }
        for day in days
    ]

    rows = []

    for user in users:
        user_code = str(
            getattr(user, "user_id", None)
            or getattr(user, "uid", None)
            or ""
        ).strip()
        assignment_key = make_assignment_key(
            getattr(user, "device_id", None),
            user_code,
        )

        for hour_label in hour_labels:
            cells = {}
            incident_cells = {}

            for day in days:
                date_key = day.isoformat()
                map_key = (assignment_key, date_key, hour_label)

                values = list(
                    dict.fromkeys(
                        attendance_by_assignment_date_hour.get(map_key, [])
                    )
                )
                incident_list = incidents_by_assignment_date_hour.get(map_key, [])

                if incident_list:
                    incident_cells[date_key] = [
                        incident_to_dict(incident)
                        for incident in incident_list
                    ]

                cells[date_key] = " / ".join(values)

            rows.append(
                {
                    "area": getattr(user, "area", None) or "",
                    "sucursal": (
                        getattr(user, "sucursal", None)
                        or getattr(getattr(user, "branch", None), "name", None)
                        or ""
                    ),
                    "trabajador": user.name,
                    "UID": user_code,
                    "uid": getattr(user, "uid", None),
                    "user_id": user_code,
                    "device_id": getattr(user, "device_id", None),
                    "assignment_key": assignment_key,
                    "hora": hour_label,
                    "empresa": getattr(user, "empresa", None) or "",
                    "cells": cells,
                    "incidents": incident_cells,
                }
            )

    return {
        "days": day_payload,
        "hours": hour_labels,
        "rows": rows,
    }


def build_payroll_excel_fallback(title: str, columns: list[str], rows_data: list[dict]) -> bytes:
    """Genera un archivo de prenómina en Excel mediante el constructor alternativo.

    :param title: Título que se mostrará en la hoja.
    :type title: str
    :param columns: Nombres de las columnas del reporte.
    :type columns: list[str]
    :param rows_data: Filas que se escribirán en el archivo.
    :type rows_data: list[dict]
    :return: Contenido binario del archivo Excel generado.
    :rtype: bytes
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Prenomina"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    header_fill = PatternFill("solid", fgColor="FCE4D6")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_index, column in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_index, value=column)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_index, row in enumerate(rows_data, start=3):
        for col_index, column in enumerate(columns, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=row.get(column, ""))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(max_length + 3, 12), 28)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def user_to_dict(user):
    """Convierte un empleado en un diccionario serializable.

    :param user: Empleado que se desea convertir.
    :type user: object
    :return: Diccionario con los datos disponibles del empleado.
    :rtype: dict
    """
    return {
        "id": user.id,
        "uid": user.uid,
        "user_id": user.user_id,
        "name": user.name,
        "role": user.role.value if hasattr(user.role, "value") else str(user.role),
        "status": user.status if hasattr(user, "status") else "Activo",
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "updated_at": user.updated_at.isoformat() if user.updated_at else None,
        "sucursal": user.sucursal if hasattr(user, "sucursal") else None,
        "email": user.email if hasattr(user, "email") else None,
        "area": user.area if hasattr(user, "area") else None,
        "empresa": user.empresa if hasattr(user, "empresa") else None,
        "branch_id": getattr(user, "branch_id", None),
        "device_id": getattr(user, "device_id", None),
        "device_code": getattr(user, "device_code", None),
    }


def attendance_to_dict(record):
    """Convierte un registro de asistencia en un diccionario serializable.

    :param record: Registro de asistencia que se desea convertir.
    :type record: object
    :return: Diccionario con los datos disponibles de la asistencia.
    :rtype: dict
    """
    return {
        "id": record.id,
        "uid": record.uid,
        "user_id": record.user_id,
        "name": record.name,
        "timestamp": record.timestamp.isoformat() if record.timestamp else None,
        "status": record.status,
        "synced_at": record.synced_at.isoformat() if getattr(record, "synced_at", None) else None,
        "branch_id": getattr(record, "branch_id", None),
        "device_id": getattr(record, "device_id", None),
        "device_code": getattr(record, "device_code", None),
    }


def attendance_to_excel_dict(record):
    """Convierte un registro de asistencia al formato utilizado por el exportador de Excel.

    :param record: Registro de asistencia que se desea convertir.
    :type record: object
    :return: Diccionario con los campos requeridos para el archivo Excel.
    :rtype: dict
    """
    return {
        "uid": record.uid,
        "user_id": record.user_id,
        "name": record.name,
        "timestamp": record.timestamp,
        "status": record.status,
    }


def utc_iso(value):
    """Serializa una fecha UTC en formato ISO para su conversión en el navegador.

    :param value: Fecha u hora que se desea serializar.
    :type value: datetime | None
    :return: Cadena ISO terminada en ``Z`` o ``None`` cuando no existe un valor.
    :rtype: str | None
    """
    if value is None:
        return None

    iso_value = value.isoformat()
    return iso_value if iso_value.endswith("Z") else f"{iso_value}Z"


def device_to_dict(device):
    """Convierte un reloj biométrico y su estado de sincronización en un diccionario.

    :param device: Reloj biométrico que se desea convertir.
    :type device: object
    :return: Diccionario con la configuración y el estado disponible del reloj.
    :rtype: dict
    """
    sync_state = ZKService.get_sync_state(device.ip, device.port)
    return {
        "id": device.id,
        "nombre": device.name,
        "name": device.name,
        "ip": device.ip,
        "ip_address": device.ip,
        "puerto": device.port,
        "port": device.port,
        "sucursal": device.location,
        "ubicacion": device.description,
        "location": device.location,
        "description": device.description,
        "empresa": getattr(device, "empresa", None),
        "password": getattr(device, "password", ""),
        "device_password": getattr(device, "password", ""),
        "activo": device.is_active,
        "is_active": device.is_active,
        "estado": device.status,
        "status": device.status,
        "sync_in_progress": bool(sync_state and sync_state.get("active")),
        "sync_connected": sync_state.get("connected") if sync_state else None,
        "sync_cancelled": bool(sync_state and sync_state.get("cancelled")),
        "sync_error": sync_state.get("reason") if sync_state else None,
        "branch_id": getattr(device, "branch_id", None),
        "ultima_sincronizacion": utc_iso(device.last_sync_at or device.last_connection),
        "last_sync_at": utc_iso(getattr(device, "last_sync_at", None)),
        "auto_sync_enabled": bool(getattr(device, "auto_sync_enabled", True)),
        "sync_interval_minutes": int(getattr(device, "sync_interval_minutes", 4) or 4),
        "next_sync_at": utc_iso(
            device.last_sync_at
            + timedelta(minutes=int(getattr(device, "sync_interval_minutes", 4) or 4))
            if bool(getattr(device, "auto_sync_enabled", True))
            and getattr(device, "last_sync_at", None)
            else None
        ),
        "created_at": utc_iso(device.created_at),
        "updated_at": utc_iso(device.updated_at),
    }


def test_device_connection(ip: str, port: int, timeout: float = 2.0) -> bool:
    """Comprueba el estado de conexión de un reloj biométrico.

    :param ip: Dirección IP del reloj.
    :type ip: str
    :param port: Puerto de comunicación del reloj.
    :type port: int
    :param timeout: Tiempo máximo de espera para la comprobación.
    :type timeout: float
    :return: ``True`` si el reloj responde como conectado; en caso contrario, ``False``.
    :rtype: bool
    """
    # Usa el mismo candado de ZKService para no abrir sondeos TCP mientras el
    # reloj está descargando usuarios o asistencias.
    return ZKService.check_device_status(
        ip=ip,
        port=int(port),
        timeout=max(1, int(timeout)),
    )


def update_device_status_safely(device, status: str):
    """Intenta actualizar el estado almacenado de un reloj sin propagar errores de actualización.

    :param device: Reloj cuyo estado se desea actualizar.
    :type device: object
    :param status: Nuevo estado que se desea registrar.
    :type status: str
    :return: No devuelve ningún valor.
    :rtype: None
    """
    try:
        DBService.update_device_status(device.id, status)
    except TypeError:
        try:
            DBService.update_device_status(device.id, status=status)
        except Exception as err:
            print(f"[WARN] No se pudo actualizar el estado del reloj {device.id}: {err}")
    except AttributeError:
        print("[WARN] DBService.update_device_status no existe")
    except Exception as err:
        print(f"[WARN] No se pudo actualizar el estado del reloj {device.id}: {err}")



def get_branch_or_404(branch_id: int):
    """Obtiene una sucursal por su identificador o genera una respuesta de recurso no encontrado.

    :param branch_id: Identificador interno de la sucursal.
    :type branch_id: int
    :return: Sucursal encontrada.
    :rtype: object
    :raises HTTPException: Si no existe una sucursal con el identificador proporcionado.
    """
    branch = DBService.get_branch_by_id(branch_id)

    if not branch:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")

    return branch


def filter_records_by_range(records, start_datetime: datetime, end_datetime: datetime):
    """Filtra registros de asistencia por un intervalo de fecha y hora inclusivo.

    :param records: Registros que se desean filtrar.
    :type records: iterable
    :param start_datetime: Límite inicial del intervalo.
    :type start_datetime: datetime
    :param end_datetime: Límite final del intervalo.
    :type end_datetime: datetime
    :return: Lista de registros cuyo sello de tiempo pertenece al intervalo.
    :rtype: list
    """
    return [
        record
        for record in records
        if record.timestamp
        and record.timestamp >= start_datetime
        and record.timestamp <= end_datetime
    ]


def get_attendance_records(
    start_datetime: datetime = datetime.min,
    end_datetime: datetime = datetime.max,
    branch_id: Optional[int] = None,
):
    """Obtiene registros de asistencia dentro de un intervalo y, opcionalmente, de una sucursal.

    :param start_datetime: Fecha y hora inicial del intervalo.
    :type start_datetime: datetime
    :param end_datetime: Fecha y hora final del intervalo.
    :type end_datetime: datetime
    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :return: Registros de asistencia que cumplen los filtros indicados.
    :rtype: list
    :raises HTTPException: Si se proporciona una sucursal inexistente.
    """
    if branch_id is not None:
        get_branch_or_404(branch_id)
        records = DBService.get_attendance_by_branch(branch_id)
        return filter_records_by_range(records, start_datetime, end_datetime)

    return DBService.get_attendance_by_date_range(
        start_datetime,
        end_datetime
    )


@router.get("/users/paginated", summary="Obtener usuarios paginados desde PostgreSQL")
def get_users_paginated_from_db(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    branch_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
):
    """Obtiene empleados paginados desde la base de datos.

    :param page: Número de página solicitado.
    :type page: int
    :param limit: Cantidad máxima de registros por página.
    :type limit: int
    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :param search: Texto opcional de búsqueda.
    :type search: Optional[str]
    :param status: Estado opcional para filtrar empleados.
    :type status: Optional[str]
    :return: Respuesta con los empleados y la información de paginación.
    :rtype: dict
    :raises HTTPException: Si se proporciona una sucursal inexistente.
    """
    if branch_id is not None:
        get_branch_or_404(branch_id)

    result = DBService.get_users_paginated(
        page=page,
        limit=limit,
        branch_id=branch_id,
        search=search,
        status=status,
    )
    data = [user_to_dict(user) for user in result["items"]]
    total = result["total"]

    return success(
        data={
            "items": data,
            "page": page,
            "limit": limit,
            "total": total,
            "pages": max(1, (total + limit - 1) // limit),
        },
        message=f"Se obtuvieron {len(data)} de {total} usuarios",
    )


@router.get("/users", summary="Obtener usuarios desde PostgreSQL")
def get_users_from_db(
    branch_id: Optional[int] = Query(
        None,
        description="ID de sucursal para filtrar empleados"
    )
):
    """Obtiene empleados desde la base de datos y permite filtrarlos por sucursal.

    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :return: Respuesta con la lista de empleados encontrados.
    :rtype: dict
    :raises HTTPException: Si se proporciona una sucursal inexistente.
    """
    if branch_id is not None:
        get_branch_or_404(branch_id)
        users = DBService.get_users_by_branch(branch_id)
    else:
        users = DBService.get_all_users_from_db()

    data = [user_to_dict(user) for user in users]

    return success(
        data=data,
        message=f"Se obtuvieron {len(data)} usuarios desde la base de datos"
    )


@router.get("/attendance/paginated", summary="Obtener asistencias paginadas desde PostgreSQL")
def get_attendance_paginated_from_db(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    branch_id: Optional[int] = Query(None),
    user_ids: Optional[str] = Query(None),
):
    """Obtiene asistencias paginadas aplicando los filtros proporcionados.

    :param page: Número de página solicitado.
    :type page: int
    :param limit: Cantidad máxima de registros por página.
    :type limit: int
    :param start_date: Fecha inicial opcional del filtro.
    :type start_date: Optional[str]
    :param end_date: Fecha final opcional del filtro.
    :type end_date: Optional[str]
    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :param user_ids: Códigos opcionales de empleados separados por comas.
    :type user_ids: Optional[str]
    :return: Respuesta con las asistencias y la información de paginación.
    :rtype: dict
    :raises HTTPException: Si una fecha es inválida o la sucursal no existe.
    """
    if branch_id is not None:
        get_branch_or_404(branch_id)

    start = datetime.combine(parse_date(start_date), time.min) if start_date else None
    end = datetime.combine(parse_date(end_date), time.max) if end_date else None

    result = DBService.get_attendance_paginated(
        page=page,
        limit=limit,
        start_date=start,
        end_date=end,
        branch_id=branch_id,
        user_ids=parse_user_ids(user_ids),
    )
    data = [attendance_to_dict(record) for record in result["items"]]
    total = result["total"]

    return success(
        data={
            "items": data,
            "page": page,
            "limit": limit,
            "total": total,
            "pages": max(1, (total + limit - 1) // limit),
        },
        message=f"Se obtuvieron {len(data)} de {total} asistencias",
    )


@router.get("/attendance", summary="Obtener asistencias desde PostgreSQL")
def get_attendance_from_db(
    limit: int = Query(100, ge=1, le=1000),
    branch_id: Optional[int] = Query(
        None,
        description="ID de sucursal para filtrar asistencias"
    )
):
    """Obtiene una cantidad limitada de asistencias desde la base de datos.

    :param limit: Cantidad máxima de registros que se devolverán.
    :type limit: int
    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :return: Respuesta con la lista de asistencias encontradas.
    :rtype: dict
    :raises HTTPException: Si se proporciona una sucursal inexistente.
    """
    records = get_attendance_records(branch_id=branch_id)
    records = records[:limit]

    data = [attendance_to_dict(record) for record in records]

    return success(
        data=data,
        message=f"Se obtuvieron {len(data)} asistencias desde la base de datos"
    )


@router.get("/attendance/dates", summary="Obtener fechas con asistencias desde PostgreSQL")
def get_attendance_dates_from_db(
    branch_id: Optional[int] = Query(
        None,
        description="ID de sucursal para filtrar fechas con asistencia"
    )
):
    """Obtiene las fechas que contienen registros de asistencia.

    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :return: Respuesta con las fechas y el total de registros por fecha.
    :rtype: dict
    :raises HTTPException: Si se proporciona una sucursal inexistente.
    """
    if branch_id is None:
        dates = DBService.get_attendance_dates_summary()

        return success(
            data=dates,
            message=f"Se obtuvieron {len(dates)} fechas con registros"
        )

    get_branch_or_404(branch_id)
    records = DBService.get_attendance_by_branch(branch_id)

    grouped = {}

    for record in records:
        if not record.timestamp:
            continue

        fecha = record.timestamp.date().isoformat()
        grouped[fecha] = grouped.get(fecha, 0) + 1

    dates = [
        {
            "fecha": fecha,
            "total": total,
        }
        for fecha, total in sorted(grouped.items(), reverse=True)
    ]

    return success(
        data=dates,
        message=f"Se obtuvieron {len(dates)} fechas con registros"
    )


@router.get("/attendance/report", summary="Obtener reporte de asistencias desde PostgreSQL")
def get_attendance_report_from_db(
    start_date: str = Query(..., description="Fecha inicial YYYY-MM-DD, DD-MM-YYYY o DD/MM/YYYY"),
    end_date: str = Query(..., description="Fecha final YYYY-MM-DD, DD-MM-YYYY o DD/MM/YYYY"),
    branch_id: Optional[int] = Query(
        None,
        description="ID de sucursal para filtrar reporte"
    )
):
    """Obtiene un reporte de asistencias para el intervalo indicado.

    :param start_date: Fecha inicial del reporte.
    :type start_date: str
    :param end_date: Fecha final del reporte.
    :type end_date: str
    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :return: Respuesta con los registros de asistencia del periodo.
    :rtype: dict
    :raises HTTPException: Si una fecha es inválida o la sucursal no existe.
    """
    start = datetime.combine(parse_date(start_date), time.min)
    end = datetime.combine(parse_date(end_date), time.max)

    records = get_attendance_records(
        start_datetime=start,
        end_datetime=end,
        branch_id=branch_id,
    )

    data = [attendance_to_dict(record) for record in records]

    return success(
        data=data,
        message=f"Se obtuvieron {len(data)} registros desde {start_date} hasta {end_date}"
    )


@router.get("/attendance/today", summary="Obtener asistencias del día actual")
def get_today_attendance(
    branch_id: Optional[int] = Query(
        None,
        description="ID de sucursal para filtrar asistencias de hoy"
    )
):
    """Obtiene las asistencias correspondientes al día actual.

    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :return: Respuesta con las asistencias del día actual.
    :rtype: dict
    :raises HTTPException: Si se proporciona una sucursal inexistente.
    """
    today = datetime.now().date()

    start_datetime = datetime.combine(today, time.min)
    end_datetime = datetime.combine(today, time.max)

    records = get_attendance_records(
        start_datetime=start_datetime,
        end_datetime=end_datetime,
        branch_id=branch_id,
    )

    data = [attendance_to_dict(record) for record in records]

    return success(
        data=data,
        message=f"Se obtuvieron {len(data)} asistencias de hoy"
    )


@router.get("/attendance/week", summary="Obtener asistencias de esta semana")
def get_week_attendance(
    branch_id: Optional[int] = Query(
        None,
        description="ID de sucursal para filtrar asistencias de la semana"
    )
):
    """Obtiene las asistencias correspondientes a la semana actual.

    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :return: Respuesta con las asistencias de la semana actual.
    :rtype: dict
    :raises HTTPException: Si se proporciona una sucursal inexistente.
    """
    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)

    start_datetime = datetime.combine(start_of_week, time.min)
    end_datetime = datetime.combine(end_of_week, time.max)

    records = get_attendance_records(
        start_datetime=start_datetime,
        end_datetime=end_datetime,
        branch_id=branch_id,
    )

    data = [attendance_to_dict(record) for record in records]

    return success(
        data=data,
        message=f"Se obtuvieron {len(data)} asistencias de esta semana"
    )


class PayrollIncidentCreate(BaseModel):
    """Representa los datos utilizados para crear o actualizar una incidencia de prenómina."""
    id: Optional[int] = None
    device_id: int
    user_id: str
    fecha: str
    hora: Optional[str] = None
    incidencia: str
    descripcion: Optional[str] = None
    color: str = "#BAE6FD"


@router.get("/prenomina", summary="Obtener prenómina con asistencias e incidencias")
@router.get("/payroll/report", summary="Obtener prenómina con asistencias e incidencias")
def get_payroll_report(
    start_date: str = Query(..., description="Fecha inicial YYYY-MM-DD"),
    end_date: str = Query(..., description="Fecha final YYYY-MM-DD"),
    branch_id: Optional[int] = Query(
        None,
        description="ID de sucursal para filtrar prenómina"
    ),
    user_ids: Optional[str] = Query(
        None,
        description="Códigos de empleados separados por coma"
    ),
):
    """Genera la respuesta de prenómina con asistencias e incidencias del periodo indicado.

    :param start_date: Fecha inicial del periodo.
    :type start_date: str
    :param end_date: Fecha final del periodo.
    :type end_date: str
    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :param user_ids: Códigos o claves opcionales de empleados separados por comas.
    :type user_ids: Optional[str]
    :return: Respuesta con la estructura de prenómina generada.
    :rtype: dict
    :raises HTTPException: Si los filtros o los servicios requeridos no son válidos.
    """
    data = build_payroll_data(
        start_date=start_date,
        end_date=end_date,
        branch_id=branch_id,
        user_ids=parse_user_ids(user_ids),
    )

    return success(
        data=data,
        message=f"Se generó la prenómina de {start_date} a {end_date}"
    )


@router.post("/prenomina/incidencias", summary="Guardar incidencia de prenómina")
@router.post("/payroll/incidents", summary="Guardar incidencia de prenómina")
def save_payroll_incident(payload: PayrollIncidentCreate):
    """Guarda o actualiza una incidencia de prenómina.

    :param payload: Datos de la incidencia que se desea guardar.
    :type payload: PayrollIncidentCreate
    :return: Respuesta con la incidencia almacenada.
    :rtype: dict
    :raises HTTPException: Si el servicio requerido no existe o los datos no pueden procesarse.
    """
    try:
        incident = DBService.save_payroll_incident(
            user_id=payload.user_id,
            fecha=parse_date(payload.fecha),
            hora=parse_hour(payload.hora) if payload.hora else None,
            incidencia=payload.incidencia,
            color=payload.color,
            incident_id=payload.id,
            device_id=payload.device_id,
        )
    except AttributeError:
        raise HTTPException(
            status_code=500,
            detail="Falta DBService.save_payroll_incident en el backend"
        )
    except Exception as err:
        raise HTTPException(status_code=400, detail=str(err))

    try:
        DBService.create_log(
            accion="Incidencia de prenómina guardada",
            detalle=(
                f"Empleado {incident.user_id} - {incident.fecha}: "
                f"{incident.incidencia}"
            )
        )
    except Exception:
        pass

    return success(
        data=incident_to_dict(incident),
        message="Incidencia guardada correctamente"
    )


@router.delete("/prenomina/incidencias/{incident_id}", summary="Inactivar incidencia de prenómina")
@router.delete("/payroll/incidents/{incident_id}", summary="Inactivar incidencia de prenómina")
def delete_payroll_incident(incident_id: int):
    """Inactiva una incidencia de prenómina sin eliminarla de la base de datos.

    :param incident_id: Identificador interno de la incidencia.
    :type incident_id: int
    :return: Respuesta con el identificador de la incidencia inactivada.
    :rtype: dict
    :raises HTTPException: Si el servicio requerido no existe o la incidencia no fue encontrada.
    """
    try:
        deleted = DBService.delete_payroll_incident(incident_id)
    except AttributeError:
        raise HTTPException(
            status_code=500,
            detail="Falta DBService.delete_payroll_incident en el backend"
        )

    if not deleted:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")

    return success(
        data={"id": incident_id},
        message="Incidencia inactivada correctamente"
    )


@router.get("/prenomina/download", summary="Descargar prenómina en Excel")
@router.get("/payroll/report/download", summary="Descargar prenómina en Excel")
def download_payroll_report(
    start_date: str = Query(..., description="Fecha inicial YYYY-MM-DD"),
    end_date: str = Query(..., description="Fecha final YYYY-MM-DD"),
    branch_id: Optional[int] = Query(
        None,
        description="ID de sucursal para filtrar prenómina"
    ),
    user_ids: Optional[str] = Query(
        None,
        description="Códigos de empleados separados por coma"
    ),
):
    """Genera y descarga la prenómina del periodo en formato Excel.

    :param start_date: Fecha inicial del periodo.
    :type start_date: str
    :param end_date: Fecha final del periodo.
    :type end_date: str
    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :param user_ids: Códigos o claves opcionales de empleados separados por comas.
    :type user_ids: Optional[str]
    :return: Respuesta de descarga con el archivo Excel generado.
    :rtype: StreamingResponse
    :raises HTTPException: Si los filtros o los servicios requeridos no son válidos.
    """
    report = build_payroll_data(
        start_date=start_date,
        end_date=end_date,
        branch_id=branch_id,
        user_ids=parse_user_ids(user_ids),
    )

    columns = ["EMPRESA", "AREA", "SUCURSAL", "TRABAJADOR"]
    columns.extend(day["label"] for day in report["days"])

    grouped = {}

    for row in report["rows"]:
        key = (
            row.get("assignment_key")
            or make_assignment_key(
                row.get("device_id"),
                row.get("user_id") or row.get("UID", ""),
            )
        )
        current = grouped.setdefault(key, {
            "AREA": row.get("area", ""),
            "SUCURSAL": row.get("sucursal", ""),
            "TRABAJADOR": row.get("trabajador", ""),
            "EMPRESA": row.get("empresa", ""),
            "__cell_colors__": {},
            "__incidents__": {},
        })

        for day in report["days"]:
            date_key = day["date"]
            column = day["label"]
            value = str(row.get("cells", {}).get(date_key, "") or "").strip()
            raw_incidents = row.get("incidents", {}).get(date_key, [])
            incident_list = (
                raw_incidents
                if isinstance(raw_incidents, list)
                else [raw_incidents] if raw_incidents else []
            )

            attendance_values = []
            for item in value.split(" / "):
                item = item.strip()
                if item and item not in attendance_values:
                    attendance_values.append(item)

            if incident_list:
                incident_lines = []
                incident_colors = []

                for incident in incident_list:
                    incident_name = str(
                        incident.get("incidencia", "INCIDENCIA")
                    ).strip().upper()
                    incident_hour = str(incident.get("hora", "") or "").strip()
                    incident_lines.append(
                        f"{incident_name} ({incident_hour})"
                        if incident_hour
                        else incident_name
                    )

                    incident_color = incident.get("color") or "#BAE6FD"
                    if incident_color not in incident_colors:
                        incident_colors.append(incident_color)

                attendance_lines = [
                    f"• {item}" for item in attendance_values
                ]
                cell_text = "\n".join([
                    *incident_lines,
                    *attendance_lines,
                ])

                # Una celda de Excel sólo admite un color de fondo. Si todas
                # las incidencias usan el mismo, se conserva; si son distintos,
                # se usa un gris neutro y se muestran todas en forma de lista.
                current["__cell_colors__"][column] = (
                    incident_colors[0]
                    if len(incident_colors) == 1
                    else "#E5E7EB"
                )
                current["__incidents__"][column] = True
            else:
                cell_text = "\n".join(attendance_values)

            if cell_text:
                existing = str(current.get(column, "") or "").strip()
                if not existing:
                    current[column] = cell_text
                elif cell_text not in existing:
                    current[column] = f"{existing}\n{cell_text}"

    rows = list(grouped.values())

    builder = build_payroll_excel or build_payroll_excel_fallback
    excel_bytes = builder(
        title="PERIODO DE ASISTENCIA PRENÓMINA EMPRESA",
        columns=columns,
        rows_data=rows,
    )

    filename = f"prenomina_{start_date.replace('/', '-')}_a_{end_date.replace('/', '-')}.xlsx"

    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


class DeviceCreate(BaseModel):
    """Representa los datos requeridos para registrar un reloj biométrico."""
    nombre: str
    ip: str
    password: str
    puerto: int = 4370
    sucursal: Optional[str] = None
    ubicacion: Optional[str] = None
    empresa: Optional[EmpresaDevice] = "FISMAN"
    branch_id: Optional[int] = None
    auto_sync_enabled: bool = True
    sync_interval_minutes: int = 4


class DeviceUpdate(BaseModel):
    """Representa los campos opcionales utilizados para actualizar un reloj biométrico."""
    nombre: Optional[str] = None
    ip: Optional[str] = None
    password: Optional[str] = None
    puerto: Optional[int] = None
    sucursal: Optional[str] = None
    ubicacion: Optional[str] = None
    empresa: Optional[EmpresaDevice] = None
    activo: Optional[bool] = None
    branch_id: Optional[int] = None
    auto_sync_enabled: Optional[bool] = None
    sync_interval_minutes: Optional[int] = None


@router.get("/devices", summary="Obtener relojes registrados desde PostgreSQL")
def get_devices_from_db(
    branch_id: Optional[int] = Query(
        None,
        description="ID de sucursal para filtrar relojes"
    )
):
    """Obtiene los relojes biométricos registrados y permite filtrarlos por sucursal.

    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :return: Respuesta con la lista de relojes registrados.
    :rtype: dict
    :raises HTTPException: Si se proporciona una sucursal inexistente.
    """
    if branch_id is not None:
        get_branch_or_404(branch_id)
        devices = DBService.get_devices_by_branch(branch_id)
    else:
        devices = DBService.get_all_devices()

    data = [device_to_dict(device) for device in devices]

    return success(
        data=data,
        message=f"Se obtuvieron {len(data)} relojes registrados"
    )


@router.post("/devices/check-status", summary="Verificar estado de conexión de relojes")
def check_devices_status(
    branch_id: Optional[int] = Query(
        None,
        description="ID de sucursal para filtrar relojes"
    )
):
    """Verifica el estado de conexión de los relojes biométricos registrados.

    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :return: Respuesta con el estado actualizado de los relojes verificados.
    :rtype: dict
    :raises HTTPException: Si se proporciona una sucursal inexistente.
    """
    if branch_id is not None:
        get_branch_or_404(branch_id)
        devices = DBService.get_devices_by_branch(branch_id)
    else:
        devices = DBService.get_all_devices()

    data = []

    for device in devices:
        is_active = bool(getattr(device, "is_active", True))

        if not is_active:
            status = "Inactivo"
        else:
            sync_state = ZKService.get_sync_state(device.ip, device.port)
            if sync_state is not None:
                connected = bool(sync_state.get("connected"))
            else:
                connected = test_device_connection(device.ip, device.port)
            # Una desconexión detectada durante sincronización queda fijada
            # hasta confirmar recuperación con dos sondeos reales. Esto también
            # descarta resultados atrasados que terminaron antes del watchdog.
            if connected and ZKService.is_disconnect_latched(device.ip, device.port):
                connected = False
            status = "Conectado" if connected else "Desconectado"

        update_device_status_safely(device, status)

        device.status = status

        if status == "Conectado":
            device.last_connection = datetime.now()

        data.append(device_to_dict(device))

    return success(
        data=data,
        message=f"Se verificaron {len(data)} relojes"
    )



@router.post("/devices", summary="Registrar reloj biométrico en PostgreSQL")
def create_device(device: DeviceCreate):
    """Registra un reloj biométrico en la base de datos.

    :param device: Datos de configuración del reloj que se desea registrar.
    :type device: DeviceCreate
    :return: Respuesta con el reloj registrado.
    :rtype: dict
    :raises HTTPException: Si el intervalo de sincronización, la contraseña o la sucursal no son válidos.
    """
    if device.sync_interval_minutes < 1 or device.sync_interval_minutes > 60:
        raise HTTPException(status_code=400, detail="El intervalo debe estar entre 1 y 60 minutos")

    if not device.password or not device.password.strip():
        raise HTTPException(status_code=400, detail="La contraseña del reloj es obligatoria")

    sucursal = device.sucursal

    if device.branch_id is not None:
        branch = get_branch_or_404(device.branch_id)
        sucursal = branch.name

    saved = DBService.save_device(
        nombre=device.nombre,
        ip=device.ip,
        puerto=device.puerto,
        password=device.password.strip(),
        sucursal=sucursal,
        ubicacion=device.ubicacion,
        empresa=device.empresa,
        branch_id=device.branch_id,
        auto_sync_enabled=device.auto_sync_enabled,
        sync_interval_minutes=device.sync_interval_minutes,
    )

    DBService.create_log(
        accion="Reloj agregado",
        detalle=f"Se registró el reloj {saved.name} ({saved.ip})"
    )

    return success(
        data=device_to_dict(saved),
        message="Reloj registrado correctamente"
    )


@router.get("/devices/{device_id}", summary="Obtener reloj por ID desde PostgreSQL")
def get_device_by_id(device_id: int):
    """Obtiene un reloj biométrico por su identificador interno.

    :param device_id: Identificador interno del reloj.
    :type device_id: int
    :return: Respuesta con los datos del reloj encontrado.
    :rtype: dict
    :raises HTTPException: Si el reloj no existe.
    """
    device = DBService.get_device_by_id(device_id)

    if not device:
        raise HTTPException(status_code=404, detail="Reloj no encontrado")

    return success(
        data=device_to_dict(device),
        message="Reloj obtenido correctamente"
    )


@router.get(
    "/devices/{device_id}/time-status",
    summary="Comparar fecha y hora del reloj con el servidor",
)
def get_device_time_status(device_id: int):
    """Compara la fecha y hora de un reloj biométrico con las del servidor.

    :param device_id: Identificador interno del reloj.
    :type device_id: int
    :return: Respuesta con el estado de sincronización de fecha y hora.
    :rtype: dict
    :raises HTTPException: Si el reloj no existe.
    """
    device = DBService.get_device_by_id(device_id)

    if not device:
        raise HTTPException(status_code=404, detail="Reloj no encontrado")

    clock_status = ZKService.get_device_time_status(
        ip=device.ip,
        port=device.port,
        password=getattr(device, "password", ""),
    )

    return success(
        data={
            "device_id": device.id,
            "device_name": device.name,
            "ip": device.ip,
            **clock_status,
        },
        message=(
            "La fecha y hora del reloj son correctas"
            if clock_status["in_sync"]
            else "La fecha y hora del reloj están desfasadas"
        ),
    )


@router.post(
    "/devices/{device_id}/sync-time",
    summary="Ajustar fecha y hora del reloj con la hora del servidor",
)
def sync_device_time(device_id: int):
    """Ajusta la fecha y hora de un reloj biométrico con la hora del servidor.

    :param device_id: Identificador interno del reloj.
    :type device_id: int
    :return: Respuesta con el resultado del ajuste de fecha y hora.
    :rtype: dict
    :raises HTTPException: Si el reloj no existe.
    :raises DeviceClockDriftError: Si el reloj no confirma correctamente el ajuste.
    """
    device = DBService.get_device_by_id(device_id)

    if not device:
        raise HTTPException(status_code=404, detail="Reloj no encontrado")

    clock_status = ZKService.sync_device_time(
        ip=device.ip,
        port=device.port,
        password=getattr(device, "password", ""),
    )

    if not clock_status.get("in_sync", False):
        raise DeviceClockDriftError(
            message="El reloj no confirmó el ajuste de fecha y hora.",
            details=clock_status,
        )

    DBService.create_log(
        accion="Hora de reloj actualizada",
        detalle=(
            f"Se ajustó la fecha y hora de {device.name} ({device.ip}) "
            f"con la hora del servidor"
        ),
    )

    return success(
        data={
            "device_id": device.id,
            "device_name": device.name,
            "ip": device.ip,
            **clock_status,
        },
        message="Fecha y hora del reloj actualizadas correctamente",
    )


@router.put("/devices/{device_id}", summary="Actualizar reloj biométrico")
def update_device(device_id: int, device: DeviceUpdate):
    """Actualiza la configuración de un reloj biométrico.

    :param device_id: Identificador interno del reloj.
    :type device_id: int
    :param device: Campos que se desean actualizar.
    :type device: DeviceUpdate
    :return: Respuesta con los datos actualizados del reloj.
    :rtype: dict
    :raises HTTPException: Si los datos son inválidos, la sucursal no existe o el reloj no fue encontrado.
    """
    if device.sync_interval_minutes is not None and not 1 <= device.sync_interval_minutes <= 60:
        raise HTTPException(status_code=400, detail="El intervalo debe estar entre 1 y 60 minutos")

    if device.password is not None and not device.password.strip():
        raise HTTPException(status_code=400, detail="La contraseña del reloj no puede quedar vacía")

    sucursal = device.sucursal

    if device.branch_id is not None:
        branch = get_branch_or_404(device.branch_id)
        sucursal = branch.name

    updated = DBService.update_device(
        device_id=device_id,
        nombre=device.nombre,
        ip=device.ip,
        puerto=device.puerto,
        password=device.password.strip() if device.password is not None else None,
        sucursal=sucursal,
        ubicacion=device.ubicacion,
        empresa=device.empresa,
        activo=device.activo,
        branch_id=device.branch_id,
        auto_sync_enabled=device.auto_sync_enabled,
        sync_interval_minutes=device.sync_interval_minutes,
    )

    if not updated:
        raise HTTPException(status_code=404, detail="Reloj no encontrado")

    DBService.create_log(
        accion="Reloj actualizado",
        detalle=f"Se actualizó el reloj {updated.name} ({updated.ip})"
    )

    return success(
        data=device_to_dict(updated),
        message="Reloj actualizado correctamente"
    )


@router.delete("/devices/{device_id}", summary="Inactivar reloj biométrico")
def delete_device(device_id: int):
    """Inactiva un reloj biométrico registrado.

    :param device_id: Identificador interno del reloj.
    :type device_id: int
    :return: Respuesta con el identificador del reloj inactivado.
    :rtype: dict
    :raises HTTPException: Si el reloj no existe o no puede inactivarse.
    """
    device = DBService.get_device_by_id(device_id)

    if not device:
        raise HTTPException(status_code=404, detail="Reloj no encontrado")

    device_name = device.name
    device_ip = device.ip

    deleted = DBService.delete_device(device_id)

    if not deleted:
        raise HTTPException(status_code=404, detail="Reloj no encontrado")

    DBService.create_log(
        accion="Reloj inactivado",
        detalle=f"Se inactivó el reloj {device_name} ({device_ip})"
    )

    return success(
        data={"id": device_id},
        message="Reloj inactivado correctamente"
    )


@router.put("/devices/{device_id}/activate", summary="Activar reloj biométrico")
def activate_device(device_id: int):
    """Activa un reloj biométrico registrado.

    :param device_id: Identificador interno del reloj.
    :type device_id: int
    :return: Respuesta con el identificador del reloj activado.
    :rtype: dict
    :raises HTTPException: Si el reloj no existe o no puede activarse.
    """
    device = DBService.get_device_by_id(device_id)

    if not device:
        raise HTTPException(status_code=404, detail="Reloj no encontrado")

    activated = DBService.activate_device(device_id)

    if not activated:
        raise HTTPException(status_code=404, detail="Reloj no encontrado")

    DBService.create_log(
        accion="Reloj activado",
        detalle=f"Se activó el reloj {device.name} ({device.ip})"
    )

    return success(
        data={"id": device_id},
        message="Reloj activado correctamente"
    )


@router.get("/attendance/download", summary="Descargar asistencias desde PostgreSQL en Excel")
def download_attendance_from_db(
    branch_id: Optional[int] = Query(
        None,
        description="ID de sucursal para filtrar descarga"
    )
):
    """Genera y descarga todas las asistencias disponibles en formato Excel.

    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :return: Respuesta de descarga con el archivo Excel generado.
    :rtype: StreamingResponse
    :raises HTTPException: Si se proporciona una sucursal inexistente.
    """
    records = get_attendance_records(branch_id=branch_id)

    records_dict = [attendance_to_excel_dict(record) for record in records]

    excel_bytes = build_attendance_excel(records_dict)

    filename = "asistencias_bd.xlsx"

    if branch_id is not None:
        branch = get_branch_or_404(branch_id)
        filename = f"asistencias_{branch.name.replace(' ', '_')}.xlsx"

    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/attendance/report/download", summary="Descargar reporte de asistencias desde PostgreSQL en Excel")
def download_attendance_report_from_db(
    start_date: str = Query(..., description="Fecha inicial YYYY-MM-DD, DD-MM-YYYY o DD/MM/YYYY"),
    end_date: str = Query(..., description="Fecha final YYYY-MM-DD, DD-MM-YYYY o DD/MM/YYYY"),
    branch_id: Optional[int] = Query(
        None,
        description="ID de sucursal para filtrar descarga"
    )
):
    """Genera y descarga un reporte de asistencias por periodo en formato Excel.

    :param start_date: Fecha inicial del reporte.
    :type start_date: str
    :param end_date: Fecha final del reporte.
    :type end_date: str
    :param branch_id: Identificador opcional de la sucursal.
    :type branch_id: Optional[int]
    :return: Respuesta de descarga con el archivo Excel generado.
    :rtype: StreamingResponse
    :raises HTTPException: Si una fecha es inválida o la sucursal no existe.
    """
    start = datetime.combine(parse_date(start_date), time.min)
    end = datetime.combine(parse_date(end_date), time.max)

    records = get_attendance_records(
        start_datetime=start,
        end_datetime=end,
        branch_id=branch_id,
    )

    records_dict = [attendance_to_excel_dict(record) for record in records]

    excel_bytes = build_attendance_excel(records_dict)

    filename = f"asistencias_{start_date.replace('/', '-')}_a_{end_date.replace('/', '-')}.xlsx"

    if branch_id is not None:
        branch = get_branch_or_404(branch_id)
        filename = (
            f"asistencias_{branch.name.replace(' ', '_')}_"
            f"{start_date.replace('/', '-')}_a_{end_date.replace('/', '-')}.xlsx"
        )

    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


class UserStatusUpdate(BaseModel):
    """Representa el nuevo estado que se asignará a un empleado."""
    status: str


@router.put("/users/by-id/{user_id}/status", summary="Actualizar estado por ID interno")
def update_user_status_by_id(user_id: int, payload: UserStatusUpdate):
    """Actualiza el estado de un empleado mediante su identificador interno.

    :param user_id: Identificador interno del empleado.
    :type user_id: int
    :param payload: Nuevo estado que se desea asignar.
    :type payload: UserStatusUpdate
    :return: Respuesta con los datos actualizados del empleado.
    :rtype: dict
    :raises HTTPException: Si el estado es inválido o el empleado no existe.
    """
    allowed = ["Activo", "Inactivo", "Baja"]
    if payload.status not in allowed:
        raise HTTPException(
            status_code=400,
            detail="Estado inválido. Usa Activo, Inactivo o Baja",
        )

    user = DBService.update_user_status_by_id(user_id, payload.status)
    if not user:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")

    return success(
        data=user_to_dict(user),
        message="Estado de empleado actualizado correctamente",
    )


@router.put("/users/{uid}/status", summary="Actualizar estado de empleado")
def update_user_status(uid: int, payload: UserStatusUpdate):
    """Actualiza el estado de un empleado mediante su UID.

    :param uid: UID del empleado.
    :type uid: int
    :param payload: Nuevo estado que se desea asignar.
    :type payload: UserStatusUpdate
    :return: Respuesta con los datos principales del empleado actualizado.
    :rtype: dict
    :raises HTTPException: Si el estado es inválido o el empleado no existe.
    """
    allowed = ["Activo", "Inactivo", "Baja"]

    if payload.status not in allowed:
        raise HTTPException(
            status_code=400,
            detail="Estado inválido. Usa Activo, Inactivo o Baja"
        )

    user = DBService.update_user_status(uid, payload.status)

    if not user:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")

    DBService.create_log(
        accion="Estado de empleado actualizado",
        detalle=f"Empleado {user.name} ({user.user_id}) cambió a estado {payload.status}"
    )

    return success(
        data={
            "uid": user.uid,
            "user_id": user.user_id,
            "name": user.name,
            "status": user.status,
        },
        message="Estado de empleado actualizado correctamente"
    )


class UserProfileUpdate(BaseModel):
    """Representa los campos opcionales utilizados para actualizar el perfil de un empleado."""
    role: Optional[str] = None
    sucursal: Optional[str] = None
    email: Optional[str] = None
    area: Optional[str] = None
    empresa: Optional[str] = None
    branch_id: Optional[int] = None


@router.put("/users/by-id/{user_id}/profile", summary="Actualizar perfil por ID interno")
def update_user_profile_by_id(user_id: int, payload: UserProfileUpdate):
    """Actualiza el perfil de un empleado mediante su identificador interno.

    :param user_id: Identificador interno del empleado.
    :type user_id: int
    :param payload: Campos del perfil que se desean actualizar.
    :type payload: UserProfileUpdate
    :return: Respuesta con los datos actualizados del empleado.
    :rtype: dict
    :raises HTTPException: Si la sucursal o el empleado no existen.
    """
    sucursal = payload.sucursal
    if payload.branch_id is not None:
        branch = get_branch_or_404(payload.branch_id)
        sucursal = branch.name

    user = DBService.update_user_profile_by_id(
        user_id=user_id,
        role=payload.role,
        sucursal=sucursal,
        email=payload.email,
        area=payload.area,
        empresa=payload.empresa,
        branch_id=payload.branch_id,
    )
    if not user:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")

    return success(
        data=user_to_dict(user),
        message="Perfil de empleado actualizado correctamente",
    )


@router.put("/users/{uid}/profile", summary="Actualizar perfil de empleado")
def update_user_profile(uid: int, payload: UserProfileUpdate):
    """Actualiza el perfil de un empleado mediante su UID.

    :param uid: UID del empleado.
    :type uid: int
    :param payload: Campos del perfil que se desean actualizar.
    :type payload: UserProfileUpdate
    :return: Respuesta con los datos principales del perfil actualizado.
    :rtype: dict
    :raises HTTPException: Si la sucursal o el empleado no existen.
    """
    sucursal = payload.sucursal

    if payload.branch_id is not None:
        branch = get_branch_or_404(payload.branch_id)
        sucursal = branch.name

    user = DBService.update_user_profile(
        uid=uid,
        role=payload.role,
        sucursal=sucursal,
        email=payload.email,
        area=payload.area,
        empresa=payload.empresa,
        branch_id=payload.branch_id,
    )

    if not user:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")

    DBService.create_log(
        accion="Perfil de empleado actualizado",
        detalle=f"Empleado {user.name} ({user.user_id}) actualizado"
    )

    return success(
        data={
            "uid": user.uid,
            "user_id": user.user_id,
            "name": user.name,
            "role": user.role.value if hasattr(user.role, "value") else str(user.role),
            "sucursal": user.sucursal,
            "email": user.email,
            "area": user.area,
            "empresa": user.empresa,
            "branch_id": getattr(user, "branch_id", None),
        },
        message="Perfil de empleado actualizado correctamente"
    )